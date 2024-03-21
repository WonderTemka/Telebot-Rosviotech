import telebot # библиотека бота
from telebot import types # класс для создания батонов

from openpyxl import load_workbook # библиотека для работы с excel табличками
import datetime # библиотека для работы с временем

from docxtpl import DocxTemplate # эту библиотеку мы используем чтобы открыть и отредактировать сертификат
from spire.doc import * # я не помню, зачем нам это надо...в документации библиотеки написано, что это тоже для работы с доками

import win32com.client
import pythoncom

bot = telebot.TeleBot('7063078819:AAHclSnusLM5G0jO9YKWVwGeTXRPevph_VA')

fn_child = 'C:/PythonProg/Bot/Bot_Vuz/vuzzz/child.xlsx' # записываем расположение эксельки для записи детей
wb_child = load_workbook(fn_child) # подгружаем эксельку по вписанному адресу
ws_child = wb_child['Sheet1'] # открываем нужный для работы лист
fn_admin = 'C:/PythonProg/Bot/Bot_Vuz/vuzzz/admin.xlsx' # записываем расположение эксельки для записи детей
wb_admin = load_workbook(fn_admin) # подгружаем эксельку по вписанному адресу
ws_admin = wb_admin['Sheet1'] # открываем нужный для работы лист
 # записываем адрес сертификата (потом будем его редактировать)
doc_anketa = DocxTemplate("C:/PythonProg/Bot/Bot_Vuz/vuzzz/sert_ank.docx")
doc_master_class = DocxTemplate("C:/PythonProg/Bot/Bot_Vuz/vuzzz/sert_mk.docx") # записываем адрес сертификата (потом будем его редактировать)
ser = "сертификат_" # текстовая переменная (это будет приставка к названию новосохраненного сертификата)
doc_path_ank = "C:/PythonProg/Bot/Bot_Vuz/vuzzz/docx/ank/" # путь к папке для новосохраненных файлов в формате doc
doc_path_mk = "C:/PythonProg/Bot/Bot_Vuz/vuzzz/docx/mk/" # путь к папке для новосохраненных файлов в формате doc
pdf_path_ank = "C:/PythonProg/Bot/Bot_Vuz/vuzzz/pdf/ank/" # путь к папке для новосохраненных файлов в формате pdf (вот его мы уже отправляем детям)
pdf_path_mk = "C:/PythonProg/Bot/Bot_Vuz/vuzzz/pdf/mk/" # путь к папке для новосохраненных файлов в формате pdf (вот его мы уже отправляем детям)

questions = ['1. Вы любите смотреть фильмы и передачи из жизни живой природы?',
            '2. Вы бы смогли перебороть страх, например, перед крысой, если вам понадобилось ее удержать в руках?',
            '3. Вы любите и умеете возиться с комнатными растениями, даже засохший росточек вашими усилиями превратится в цветущий куст?',
            '4. Хорошо разбираетесь в породах животных или видах растений?',
            '5. Вы считаете, что парки в городе важнее инфраструктуры (дороги, магазины, офисные помещения)?',
            '6 .Идеальный отдых для вас — не пляж, а поход за грибами, пикник в лесу или даже вылазка в горы?',
            '1. Вы легко назовете минимум 10 компонентов, из которых состоит ваш компьютер?',
            '2. Вы неплохо справляетесь с заданиями по физике, особенно лабораторными?',
            '3. Что касается техники, вам проще показать, как с ней управляться, чем объяснить?',
            '4. В детстве часто разбирали механические игрушки, чтобы узнать, что там, внутри?',
            '5. Вы предпочитаете внимательно прочитать инструкцию, чтобы освоить новый «гаджет»?',
            '6. На уроках геометрии или черчения вам не составляет большого труда точно начертить объемный рисунок, сделать его в нескольких ракурсах?',
            '1. Как ни странно, но после всех поездок у вас больше всего остается фотографий с портретами людей?',
            '2. Вы любите книги и фильмы о культуре и истории разных стран?',
            '3. Вы считаете себя сострадательным человеком, готовы помочь даже незнакомому человеку, если понадобится?',
            '4. Вы хорошо справляетесь, когда нужно присмотреть за маленькими детьми?',
            '5. Вы знаете, как убедить человека сделать то, что вам нужно, умеете найти свой подход к каждому?',
            '6. Вы можете справиться с волнением и хорошо выступаете перед большой аудиторией?',
            '1.  Вы с удовольствием решаете сложные задачки, редко ошибаетесь в расчетах, умеете применять нужные формулы?',
            '2. Лучше всего информация усваивается в виде графиков и таблиц?',
            '3. Вы любите точность в расчетах, не допустите ошибок в тексте из-за невнимательности?',
            '4. Вы сможете относительно точно воспроизвести математическую формулу или схему, увидев ее первый раз?',
            '5. Если нужно объяснить что-то другому человеку (например, как организовать вечеринку), вам проще набросать схему или план?',
            '6. У вас наверняка идеальный порядок на столе, вы точно знаете, в каком ящике лежит нужный предмет?',
            '1. Когда вы что-то рассказываете, то стремитесь в первую очередь передать эмоции и переживания?',
            '2. Вы неплохо запоминаете внешность и мимику людей, сможете изобразить или нарисовать известного человека?',
            '3. Часто выбираете вещи или книги импульсивно, полагаясь на их внешний вид?',
            '4. Вы близко к сердцу принимаете эмоции героев художественных фильмов или книг, ярко представляете себе описанные пейзажи?',
            '5. У вас необычный стиль одежды или оформление комнаты… кому-то это не нравится, но вы все равно уверены, что индивидуальность — прежде всего?',
            '6. Увидев красивый сувенир или поделку, вы уверены, что сможете сделать своими руками такую же, а то и лучше?']
prognoz = ['1. Человек-природа\n'
           'Вы — друг всего живого на планете Земля! Именно такие как вы становятся исследователями богатого Царства живой природы, создают сотни новых культур растений, лечат и спасают животных. Возможно, в вас закрался и талант первооткрывателя?\n'
           'Психологические требования: наблюдательность, сопереживание, настойчивость, хладнокровие, готовность к самостоятельной работе в любых, даже трудных условиях\n'
           'Примеры профессий: биолог, биотехнолог, ветеринар, зоотехник, химик, врач, технолог (по отраслям) и др.\n'
           'Вам рекомендованы следующие образовательные направления:\n'
           '06.03.01 Биология\n'
           '36.03.01 Ветеринарно-санитарная экспертиза\n'
           '36.05.01 Ветеринария\n'
           '31.05.01 Лечебное дело\n'
           '18.03.01 Химическая технология\n'
           '19.03.01 Биотехнология\n'
           '19.03.02 Продукты питания из растительного сырья\n'
           '19.03.03 Продукты питания животного происхождения\n'
           '19.03.04 Технология продукции и организация.\n\n'
           'Основываясь на полученных аналитических данных Вы можете освоить весь перечень подходящих специальностей в ФГБОУ ВО «Российский биотехнологический университет (РОСБИОТЕХ)»!\n'
           'https://mgupp.ru/abitur/ ',
           '2. Человек-техника\n'
           'Технический прогресс не стоит на месте и вы — его вечный двигатель! Вы знаете, как управиться со сломанным будильником, сможете точно настроить любую технику, не боитесь по-новому взглянуть на обычные вещи и, прикрутив пару гаек, получить новое изобретение. Сложные механизмы не пугают, а вдохновляют вас. Но вы умеете использовать не только воображение, но и смекалку, чтобы найти общий язык с любой техникой.\n'
           'Психологические требования: внимательность, рациональность, развитое техническое мышление, хорошая координация движений, умение сосредотачиваться и быстро переключать внимание, наблюдательность.\n'
           'Примеры профессий: инженер, конструктор, техник, радиомеханик, электрик и др.\n'
           'Рекомендованы образовательные направления: \n'
           '15.03.02 Технологические машины и оборудование\n'
           '15.03.04 Автоматизация технологических процессов и производств\n'
           '15.03.06 Мехатроника и робототехника\n'
           '16.03.03 Холодильная, криогенная техника и системы жизнеобеспечения\n'
           '20.03.01 Техносферная безопасность\n'
           '27.03.01 Стандартизация и метрология\n'
           '27.03.02 Управление качеством\n'
           '27.03.04 Управление в технических системах.\n\n'
           'Основываясь на полученных аналитических данных Вы можете освоить весь перечень подходящих специальностей в ФГБОУ ВО «Российский биотехнологический университет (РОСБИОТЕХ)»!\n'
           'https://mgupp.ru/abitur/',
           '3. Человек-знаковая система\n'
           'Что уж греха таить, в точных науках вам нет равных! Математические расчеты, статистика, чертежи, информационные технологии… для вас это не просто цифры и графики — это увлекательный мир, полный находок и загадок. Для самого запутанного дела вы без труда создадите удобный план действий. Вы — неоценимый работник там, где требуется внимательность, усидчивость и хорошая память.\n'
           'Психологические требования: логика преобладает над эмоциями, умение концентрироваться, абстрактное мышление, способность оперировать отвлеченными понятиями и условными знаками, создавая системы, спокойный темперамент, терпение, внимательность.\n'
           'Примеры профессий: программист, геодезист, математик, экономист, бухгалтер, редактор, корректор, делопроизводитель и др.\n'
           'Рекомендованы образовательные направления: \n'
           '09.03.01 Информатика и вычислительная техника\n'
           '09.03.03 Прикладная информатика\n'
           '38.03.01 Экономика\n'
           '15.03.04 Автоматизация технологических процессов и производств\n'
           '15.03.06 Мехатроника и робототехника\n'
           '16.03.03 Холодильная, криогенная техника и системы жизнеобеспечения.\n\n'
           'Основываясь на полученных аналитических данных Вы можете освоить весь перечень подходящих специальностей в ФГБОУ ВО «Российский биотехнологический университет (РОСБИОТЕХ)»!\n'
           'https://mgupp.ru/abitur/',
           '4. Человек-художественный образ\n'
           'Вы созданы для того, чтобы приносить красоту в наш суровый мир! Скорее всего, вы уже давно чувствуете в себе творческую жилку — шила в мешке ведь не утаишь… В любом случае, ваше воображение, интуиция и нестандартный взгляд на окружающих пригодятся в создании чудесных произведений искусства. Будь это музыка, литература, театр, дизайн или архитектура — все вам под силу!\n'
           'Психологические требования: образное мышление, восприимчивость, чуткость, стремление к познанию нового и необычного, креативность, хорошая зрительная и звуковая память, умение владеть своим телом.\n'
           'Примеры профессий: актер, писатель, дизайнер, художник-оформитель, композитор, модельер, архитектор, скульптор, флорист, реставратор, гравер, ювелир, журналист, хореограф и др.\n'
           'Рекомендованы образовательные направления: \n'
           '29.03.03 Технология полиграфического и упаковочного\n'
           '43.03.01 Сервис\n'
           '40.03.01 Юриспруденция.\n\n'
           'Основываясь на полученных аналитических данных Вы можете освоить весь перечень подходящих специальностей в ФГБОУ ВО «Российский биотехнологический университет (РОСБИОТЕХ)»!\n'
           'https://mgupp.ru/abitur/',
           '5. «Человек-человек»\n'
           'Даже выбирая профессию, не связанную непосредственно с общением, вы поступите правильно, если обратите внимание на общительность и контактность. Подумайте, куда вы обращены- к людям или к себе? С кем бы вы хотели общаться — с собой или с другими? Главное содержание труда в профессиях типа «человек-человек» сводится к взаимодействию между людьми. Если не наладится это взаимодействие, значит, не наладится и работа. \n'
           'Качества, необходимые для работы с людьми: устойчивое, хорошее настроение в процессе работы с людьми, потребность в общении, способность мысленно ставить себя на место другого человека, быстро понимать намерения, помыслы, настроение людей, умение разбираться в человеческих взаимоотношениях, хорошая память (умение держать в уме имена и особенности многих людей), умение находить общий язык с различными людьми, терпение ... \n'
           'Рекомендованы образовательные направления: \n'
           '38.03.01 Экономика\n'
           '38.03.02 Менеджмент\n'
           '43.03.01 Сервис\n'
           '40.03.01 Юриспруденция\n'
           '38.05.02 Таможенное дело\n'
           '31.05.01 Лечебное дело\n\n'
           'Основываясь на полученных аналитических данных Вы можете освоить весь перечень подходящих специальностей в ФГБОУ ВО «Российский биотехнологический университет (РОСБИОТЕХ)»!\n'
           'https://mgupp.ru/abitur/']

ans = ['K', 'L', 'M', 'N', 'O'] #Обращение к столбцам, хранящии результаты по тестам
who_you = ['Человек-природа', 'Человек-техника', 'Человек-знаковая система', 'Человек-художественный образ', 'Человек-человек']
mon = ['', 'января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']

#Админка
def add_specify_children(callback):
    ws_admin['G2'] = int(ws_admin['G2'].value) + int(callback.text)
    wb_admin.save(fn_admin)  # сохрнаяем эксельку
    wb_admin.close()
    bot.send_message(callback.chat.id, f"Количество мест изменено")

def del_specify_children(callback):
    ws_admin['G2'] = int(ws_admin['G2'].value) - int(callback.text)
    wb_admin.save(fn_admin)  # сохрнаяем эксельку
    wb_admin.close()
    bot.send_message(callback.chat.id, f"Количество мест изменено")

def rename_mk(callback):
    rename = callback.text
    ws_admin['F2'] = rename
    wb_admin.save(fn_admin)  # сохрнаяем эксельку
    wb_admin.close()
    bot.send_message(callback.chat.id, f"Название изменено на: {rename}")

def age_child(callback):
    age = callback.text
    ws_admin['H2'] = age
    wb_admin.save(fn_admin)  # сохрнаяем эксельку
    wb_admin.close()
    bot.send_message(callback.chat.id, f"Возраст детей указн: {age}")

def name_school(callback):
    school = callback.text
    ws_admin['I2'] = school
    wb_admin.save(fn_admin)  # сохрнаяем эксельку
    wb_admin.close()
    bot.send_message(callback.chat.id, f"Название школы изменено: {school}")

def meropriyatie(callback):
    mer = callback.text
    if mer.lower() == "ank" or mer.lower() == "mk":
        ws_admin["J2"] = mer.lower()
        wb_admin.save(fn_admin)  # сохрнаяем эксельку
        wb_admin.close()
        bot.send_message(callback.chat.id, f"Сейчас будет проходить мероприятие: {mer}")

def check_password(message):
    if ws_admin['E2'].value == message.text:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.InlineKeyboardButton("Посмотреть количество участницков")
        btn2 = types.InlineKeyboardButton("Количество доступных мест")
        btn3 = types.InlineKeyboardButton("Добавить места")
        btn4 = types.InlineKeyboardButton("Уменьшить места")
        btn5 = types.InlineKeyboardButton("Название мастер класса")
        btn6 = types.InlineKeyboardButton("Указать название мастеркласса")
        btn7 = types.InlineKeyboardButton("Указать возраст детей")
        btn8 = types.InlineKeyboardButton("Указать школу")
        btn9 = types.InlineKeyboardButton("Указать мероприятие")
        btn10 = types.InlineKeyboardButton("Введенный возраст детей")
        btn11 =  types.InlineKeyboardButton("Введенная школа")
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10, btn11)
        current_date = datetime.date.today()
        date = f"{current_date.day}/{current_date.month}/{current_date.year}"
        ws_admin['A2'] = message.chat.id
        ws_admin['B2'] = date
        wb_child.save(fn_child)  # сохрнаяем эксельку
        wb_child.close()
        bot.send_message(message.chat.id, "Пароль подвержден!\n"
                         f"Памятка:\n"
                         f"Самое важное необходимо укахать 'Возраст детей', 'Школу', и 'Мероприятие'.\n По необходимости указать 'название мастер-класса'"
                         f"Не забывайте так же указать количество мест при мероприятие 'Мастер-класс' и так же проверять доступное количество свободных мест"
                         ,  reply_markup=markup)
    else:
        #bot.send_message(message.chat.id, "Пароль неверный!")
        bot.register_next_step_handler(message, check_password)

#функциональная часть
def conver_dc_in_pdf(ser, title, path_doc, path_pdf):
    print(path_doc + f"{ser}{title}.docx")
    pythoncom.CoInitialize()
    Application = win32com.client.Dispatch("Word.Application")
    document_file = Application.Documents.Open(path_doc + f"{ser}{title}.docx")
    document_file.ExportAsFixedFormat(path_pdf + f"{ser}{title}.pdf", 17)
    Application.Quit()

def opros(message):
    if ws_child['R' + str(Find_id(message))].value == None:
        ws_child['R' + str(Find_id(message))] = 0
        ws_child['S' + str(Find_id(message))] = 0
        ws_child['K' + str(Find_id(message))] = 0
        ws_child['L' + str(Find_id(message))] = 0
        ws_child['M' + str(Find_id(message))] = 0
        ws_child['N' + str(Find_id(message))] = 0
        ws_child['O' + str(Find_id(message))] = 0
        wb_child.save(fn_child)  # сохрнаяем эксельку
        wb_child.close()  # закрываем эксельку (БЕЗ ЭТОГО НЕ СОХРАНИТСЯ)
    cat = int(ws_child['R' + str(Find_id(message))].value)
    que = int(ws_child['S' + str(Find_id(message))].value)
    if (int(ws_child['R' + str(Find_id(message))].value) * 6 + int(ws_child['S' + str(Find_id(message))].value))  < len(questions):
        markup = types.InlineKeyboardMarkup()
        btn1 = types.InlineKeyboardButton("да", callback_data="true")
        btn2 = types.InlineKeyboardButton("нет", callback_data="false")
        markup.row(btn1, btn2)
        bot.send_message(message.chat.id, questions[(cat * 6) + que], reply_markup=markup)
        ws_child['S' + str(Find_id(message))] = int(ws_child['S' + str(Find_id(message))].value) + 1
        if int(ws_child['S' + str(Find_id(message))].value) > 6:
            ws_child['S' + str(Find_id(message))] = 1
            ws_child['R' + str(Find_id(message))] = int(ws_child['R' + str(Find_id(message))].value) + 1
        wb_child.save(fn_child)  # сохрнаяем эксельку
        wb_child.close()  # закрываем эксельку (БЕЗ ЭТОГО НЕ СОХРАНИТСЯ)

        print(f"""категория:{int(ws_child['R' + str(Find_id(message))].value)}
вопрос:{int(ws_child['S' + str(Find_id(message))].value)}""")

def Rezultat(message):
    current_date = datetime.date.today()
    for i in range(len(ans)):
        if int(ws_child[ans[i] + str(Find_id(message))].value) >= 4:
            bot.send_message(message.chat.id, prognoz[i])
            who_ser = f"{who_you[i]}_"
            print("Выводит результат...")
            if ws_child['F' + str(Find_id(message))].value == None: # если у пользователя нет отчества (составляем сертификат с двумя словами)
                name = ws_child['D' + str(Find_id(message))].value + " " + ws_child['E' + str(Find_id(message))].value
                title = str(ws_child['D'+str(Find_id(message))].value + ws_child['E'+str(Find_id(message))].value) # создаём переменную в формате ФамилияИмя
            else: # ТОЖЕ САМОЕ, ТОЛЬКО ЕСЛИ У ПОЛЬЗОВАТЕЛЯ 3 СЛОВА В ФИО А НЕ 2
                    # получаем сегодняшнюю дату
                name = ws_child['D' + str(Find_id(message))].value + " " + ws_child['E' + str(Find_id(message))].value + " " + ws_child['F' + str(Find_id(message))].value
                title = str(ws_child['D'+str(Find_id(message))].value + ws_child['E'+str(Find_id(message))].value + ws_child['F'+str(Find_id(message))].value) # создаём переменную в формате ФамилияИмя

            context = {'name': name,
                        'total': who_you[i],
                        'day': current_date.day,
                        'month': mon[current_date.month],
                        'age': current_date.year}
            doc_anketa.render(context)  # записываем в сертификат Фамилия Имя
            doc_anketa.save(doc_path_ank + f"{ser}{who_ser[i]}{title}.docx")

            print(doc_path_ank + f"{ser}{who_ser[i]}{title}.docx")
            print(f"{ser}{who_ser[i]}{title}.docx")
            pythoncom.CoInitialize()
            Application = win32com.client.Dispatch("Word.Application")
            document_file = Application.Documents.Open(doc_path_ank + f"{ser}{who_ser[i]}{title}.docx")
            document_file.ExportAsFixedFormat(pdf_path_ank + f"{ser}{who_ser[i]}{title}.pdf", 17)
            Application.Quit()

            ws_child['J' + str(Find_id(message))] = pdf_path_ank + f"{ser}{who_ser[i]}{title}.pdf"
            wb_child.save(fn_child) # СОХРАНЯЕМ
            wb_child.close() # ЗАКРЫВАЕМ

            f = open(pdf_path_ank + f"{ser}{who_ser[i]}{title}.pdf", "rb")
            bot.send_document(message.chat.id, f)
            f.close()

# функция, принимающая сообщение пользователя, и возвращающая его порядковый номер из таблички с детьми
def Find_id(message):
    count = 1 # начинаем считать с одного, потому что индекс номер 0 занимает строка с названиями столбцов

    while True: # запускаем бесконечный цикл, в котором проходим каждую строку в таблицы с детьми
        if ws_child['A' + str(count)].value != None: # проверяем клетку столбика А на заполненость
            if ws_child['A' + str(count)].value == "id": # проверка, чтобы пропустить самую первую строку
                pass
            elif ws_child['A' + str(count)].value == message.chat.id: # если id пиздюка уже есть в таблице, значит всё пучком
                return count # возвращаем номер пиздюка
        else:
            ws_child['A' + str(count)] = message.chat.id # записываем в первую пустую клетку столбика А id ребёнка
            ws_child['C' + str(count)] = str(count - 1) # записываем в столбик с порядковыми номерами порядковый номер ребёнка

            current_date = datetime.date.today() # получаем сегодняшнюю дату
            ws_child['B' + str(count)] = f'{current_date.day}.{current_date.month}.{current_date.year}' # записываем в столбик с датами сегодняшнюю дату в нужном формате

            wb_child.save(fn_child) # сохрнаяем эксельку
            wb_child.close() # закрываем эксельку (БЕЗ ЭТОГО НЕ СОХРАНИТСЯ)
            return 0

        count += 1

# функция просит ребёнка предстваиться прежде чем он сможет получить сертификат или пройти тест
def SayMyName1(message):
    if ws_child['D'+str(Find_id(message))].value == None: # если у нас нет имени пиздюка - выпрашиваем его
        bot.send_message(message.chat.id, "Прежде, чем мы продолжим, мне нужно записать твои данные😁\n"
                                    "Мне нужно твое ФИО, почта и номер телефона✍️\n"
                                    "Для начала как я могу к тебе обращаться?\n\n"
                                    "Напиши свое ФИО")
        bot.register_next_step_handler(message, SayMyName2) # после ответа пользователя переходим в функцию подтверждения имени
    else:
        Phone1(message)

# функция подтверждения имени (чтобы бот переспросил, а ребёнок подтвердил, что имя введено верно)
def SayMyName2(message):
    fio = message.text.split() # массив запоминает имя, фамилию и отчество(опционально) как разные элементы
    markup = types.InlineKeyboardMarkup()
    btn1 = types.InlineKeyboardButton("да", callback_data="да, это моё имя")
    btn2 = types.InlineKeyboardButton("нет", callback_data="нет, это не моё имя")
    markup.row(btn1, btn2)
    ws_child['E' + str(Find_id(message))] = message.text # сохраняем в столбик с именем введенное сообщение на всякий случай
    wb_child.save(fn_child) # сохраняем эксельку
    wb_child.close() # закрываем
    if len(fio) == 3: # если в ФИО попало 3 слова
        bot.send_message(message.chat.id, "Давай сверимся, правильно ли я записал😄\n\n"
                                            "Тебя зовут:\n"
                                            f"Фамилия: {message.text.split()[0]}\n"
                                            f"Имя: {message.text.split()[1]}\n"
                                            f"Отчество: {message.text.split()[2]}\n", reply_markup=markup)
    elif len(fio) == 2: # тоже самое, если в ФИО попало 2 слова
        bot.send_message(message.chat.id, "Давай сверимся, правильно ли я записал😄\n\n"
                                          "Тебя зовут:\n"
                                          f"Фамилия: {message.text.split()[0]}\n"
                                          f"Имя: {message.text.split()[1]}\n", reply_markup=markup)
    else: # если ребёнок ввёл не 2 и не 3 слова, значи ввод некоректный, надо об этом сказать, и дать ему ещё раз ввести имя
        bot.send_message(message.chat.id, "кажется, ты что-то перепутал😅\n"
                                          "формат для ввода имени: Фамилия Имя Отчество\n\n"
                                          "попробуй, пожалуйста, ещё раз😌")
        bot.register_next_step_handler(message, SayMyName2)

# Функция, выпрашивающая номерочек
def Phone1(message):
    if ws_child['G' + str(Find_id(message))].value == None: # если у нас есть имя но нет номерочка, клянчим номерочек
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        button_phone = types.KeyboardButton(text="Отправить телефон📲", request_contact=True) # специальная кноп_очка, клянчащая номерочек
        keyboard.add(button_phone)
        bot.send_message(message.chat.id, "Теперь я попрошу у тебя номер телефона📱\n"
                                          "Тогда мы сможем оставаться на связи🤙\n"
                                          "Можешь нажать на конпку там внизу👇", reply_markup=keyboard)

def message_email(message):
    if ws_child['H'+str(Find_id(message))].value == None: # если у человека нет файла
        a = types.ReplyKeyboardRemove() # создаю пустую клавиатуру чтобы убрать предыдущие кнопки
        bot.send_message(message.chat.id, "Отлично!\n"
                                          "Твой сертификат уже почти готов!😎\n"
                                          "Не мог бы ты ещё оставить мне свою почту, пока я заполняю сертификат?\n"
                                          "Так мы с тобой точно сможем оставаться на связи😁", reply_markup=a)
        bot.register_next_step_handler(message, input_email) # после ответа пользователя переходим в функцию подтверждения имени
    else:
        markup = types.InlineKeyboardMarkup()
        if ws_child['U'+str(Find_id(message))].value == "mk":
            btn1 = types.InlineKeyboardButton('Получить сертификат', callback_data='мастеркласс')
            markup.row(btn1)
            bot.send_message(message.chat.id, "Я тебя уже записал!"
                                        "Теперь ты можешь получить сертификат!\n", reply_markup=markup)
        elif ws_child['T'+str(Find_id(message))].value == "ank":
            btn1 = types.InlineKeyboardButton('Получить сертификат', callback_data='викторина')
            markup.row(btn1)
            bot.send_message(message.chat.id, "Я тебя уже записал!"
                                        "Теперь ты можешь получить сертификат!\n", reply_markup=markup)

def input_email(message):
    ws_child['H' + str(Find_id(message))] = message.text# сохраняем в столбик с почтой введенное сообщение
    wb_child.save(fn_child) # сохраняем эксельку
    wb_child.close() # закрываем
    markup = types.InlineKeyboardMarkup()
    if ws_child['U'+str(Find_id(message))].value == "mk":
        btn1 = types.InlineKeyboardButton('Получить сертификат', callback_data='мастеркласс')
        markup.row(btn1)
        bot.send_message(message.chat.id, "Чудно!🥳\n"
                                    "Вот твой сертификат!😁", reply_markup=markup)
    elif ws_child['T'+str(Find_id(message))].value == "ank":
        btn1 = types.InlineKeyboardButton('Поехали🚗', callback_data='викторина')
        markup.row(btn1)
        bot.send_message(message.chat.id, "Чудно!\n"
                                    "Давай перейдем к анкете!", reply_markup=markup)

@bot.message_handler(content_types=['contact']) # декоратор, который ловит пресылаемый номерочек
def contact(message):
    if message.contact is not None:
        a = types.ReplyKeyboardRemove() # создаю пустую клавиатуру чтобы убрать предыдущие кнопки

        print(message.contact.phone_number)
        ws_child['G' + str(Find_id(message))] = message.contact.phone_number # запоминаем номерочек в столбик номерочков
        wb_child.save(fn_child) # сохраняем
        wb_child.close() # и закрываем эксельку

        markup = types.InlineKeyboardMarkup()
        btn1 = types.InlineKeyboardButton('продолжаем!', callback_data="мастеркласс")
        if ws_child['U' + str(Find_id(message))].value == "mk":
            btn1 = types.InlineKeyboardButton('продолжаем!', callback_data="мастеркласс")
        elif ws_child['T' + str(Find_id(message))].value == "ank":
            btn1 = types.InlineKeyboardButton('продолжаем!', callback_data="викторина")
        markup.row(btn1)

        bot.send_message(message.chat.id, "Хорошо!😁", reply_markup=a)
        bot.send_message(message.chat.id, "Продожим дальше?", reply_markup=markup)

@bot.message_handler(commands=['start'])
def Start(message):

    if ws_admin['G2'].value - ws_child.max_row-1 < 2:
            bot.send_message(int(ws_admin['L2'].value), f"Количество доступных мест: {ws_admin['G2'].value - ws_child.max_row - 1}\n"
                            f"Необходимо добавить места")

    falag_mk_or_ank = ws_admin['J2'].value
    markup = types.InlineKeyboardMarkup()
    btn1 = types.InlineKeyboardButton('Приступим!', callback_data="мастеркласс")
    if falag_mk_or_ank == "mk":
        btn1 = types.InlineKeyboardButton('Приступим!', callback_data="мастеркласс")
    elif falag_mk_or_ank == "ank":
        btn1 = types.InlineKeyboardButton('Приступим!', callback_data='викторина')
    markup.row(btn1)

    if int(ws_child.max_row)-1 >= int(ws_admin['G2'].value) and falag_mk_or_ank == "mk":
        bot.send_message(message.chat.id, "Извини, но в настоящее время мест нет😅\n"
                                    "Ты можешь попросить администратора или управляющего решить этот вопрос")
    else:
        if Find_id(message) == 0: # если ребёнок ещё не находится в таблице, то он заносится в таблицу (см функцию Find_id) и он приветствуется
            if falag_mk_or_ank == "mk":
                bot.send_message(message.chat.id, "Привет! Я бот Росбиотеха.😃\n\n"
                                        "Я помогу тебе получить твой сертификат в электронном формате.\n"
                                        "Этот сертификат пригодиться при поступлении в университет/коледж!", reply_markup=markup)
            elif falag_mk_or_ank == "ank":
                bot.send_message(message.chat.id, "Привет! Я бот Росбиотеха.😃\n\n"
                                                  "Я помогу тебе пройти викторину и получить сертификат в электронном формате.\n"
                                                  "Прошу тебя отнестись к викторине со всей серьёзностью!🧐\n "
                                                  "Наша викторина направлена на определение склада твоего ума🧠\n"
                                                  "Определив, как ты мыслишь, мы подберём наиболее подходящие направления для твоего развития💪\n"
                                                  "К тому же наш сертификат пригодиться тебе при поступлении в университет/коледж!", reply_markup=markup)

        else: # если ребёнок уже есть в таблице, бот просто пересказывает, что он умеет
            bot.send_message(message.chat.id, "И снова здравствуй! Я тебя уже знаю😃.\n"
                                            "Хочешь продолжить?",reply_markup=markup)

# проверяем колюэки
@bot.callback_query_handler(func=lambda callback: True)
def CB(callback):
    bot.edit_message_reply_markup(callback.message.chat.id, callback.message.message_id, reply_markup=None)
    if callback.data == 'викторина': # если пиздюк хочет викторину
        ws_child['T'+str(Find_id(callback.message))] = "ank"
        ws_child['P'+str(Find_id(callback.message))] = ws_admin['H2'].value
        ws_child['Q'+str(Find_id(callback.message))] = ws_admin['I2'].value
        wb_child.save(fn_child)  # сохрнаяем эксельку
        wb_child.close()
        if ws_child['D'+str(Find_id(callback.message))].value == None: # если у нас нет имени пиздюка - выпрашиваем его
            SayMyName1(callback.message)
        elif ws_child['G'+str(Find_id(callback.message))].value == None: # если у нас есть имя но нет номерочка, клянчим номерочек
            Phone1(callback.message)
        elif ws_child['H'+str(Find_id(callback.message))].value == None: # если у нас есть имя и номерочек, но нет почты, кляньчим
            message_email(callback.message)
        elif ws_child['J'+str(Find_id(callback.message))].value != None: # если у человека уже есть файл
            f = open(ws_child['J'+str(Find_id(callback.message))].value, "rb")
            bot.send_document(callback.message.chat.id, f)
            f.close()
        else:
            opros(callback.message)
        pass
    elif callback.data == 'мастеркласс': # если пиздюк хочет сертификат, а не викторину
        ws_child['U'+str(Find_id(callback.message))] = "mk"
        ws_child['P'+str(Find_id(callback.message))] = ws_admin['H2'].value
        ws_child['Q'+str(Find_id(callback.message))] = ws_admin['I2'].value
        wb_child.save(fn_child)  # сохрнаяем эксельку
        wb_child.close()
        if ws_child['D'+str(Find_id(callback.message))].value == None: # если у нас нет имени пиздюка - выпрашиваем его
            SayMyName1(callback.message)
        elif ws_child['G' + str(Find_id(callback.message))].value == None: # если у нас есть имя но нет номерочка, клянчим номерочек
            Phone1(callback.message)
        elif ws_child['H'+str(Find_id(callback.message))].value == None: # если у нас есть имя и номерочек, но нет почты, кляньчим
            message_email(callback.message)
        elif ws_child['I'+str(Find_id(callback.message))].value != None: # если у человека уже есть файл
            f = open(ws_child['I'+str(Find_id(callback.message))].value, "rb")
            bot.send_document(callback.message.chat.id, f)
            f.close()
        else: # если имя, номерочек почта уже выпрошены, запускаем сертифицирование
            current_date = datetime.date.today()
            if ws_child['F' + str(Find_id(callback.message))].value == None: # если у пользователя нет отчества (составляем сертификат с двумя словами)
                title = str(ws_child['D'+str(Find_id(callback.message))].value + ws_child['E'+str(Find_id(callback.message))].value) # создаём переменную в формате ФамилияИмя
                name = ws_child['D' + str(Find_id(callback.message))].value + " " + ws_child['E' + str(Find_id(callback.message))].value
            else: # ТОЖЕ САМОЕ, ТОЛЬКО ЕСЛИ У ПОЛЬЗОВАТЕЛЯ 3 СЛОВА В ФИО А НЕ 2
                  # получаем сегодняшнюю дату
                name = ws_child['D' + str(Find_id(callback.message))].value + " " + ws_child['E' + str(Find_id(callback.message))].value + " " + ws_child['F' + str(Find_id(callback.message))].value
                title = str(ws_child['D'+str(Find_id(callback.message))].value + ws_child['E'+str(Find_id(callback.message))].value + ws_child['F'+str(Find_id(callback.message))].value) # создаём переменную в формате ФамилияИмя

            master_class = ws_admin['F2'].value
            context = {'name': name,
                        'master_class': master_class,
                        'day': current_date.day,
                        'month': mon[current_date.month],
                        'age': current_date.year}
            doc_master_class.render(context)  # записываем в сертификат Фамилия Имя
            doc_master_class.save(doc_path_mk + f"{ser}{title}.docx")

            print(doc_path_mk + f"{ser}{title}.docx")
            pythoncom.CoInitialize()
            Application = win32com.client.Dispatch("Word.Application")
            document_file = Application.Documents.Open(doc_path_mk + f"{ser}{title}.docx")
            document_file.ExportAsFixedFormat(pdf_path_mk + f"{ser}{title}.pdf", 17)
            Application.Quit()

            ws_child['I' + str(Find_id(callback.message))] = pdf_path_mk + f"{ser}{title}.pdf"
            wb_child.save(fn_child)  # сохрнаяем эксельку
            wb_child.close()  # закрываем эксельку (БЕЗ ЭТОГО НЕ СОХРАНИТСЯ)

            f = open(pdf_path_mk + f"{ser}{title}.pdf", "rb")
            bot.send_document(callback.message.chat.id, f)
            f.close()

    elif callback.data == 'да, это моё имя': # пиздюк подтвердил, что это его имя
        fio = ws_child['E'+str(Find_id(callback.message))].value # достаём из клеточки с именем ФИО (сейчас это одна строка)
        print(fio)
        fio = fio.split() # режем ФИО на Ф И О
        if len(fio) == 2: # если у нас ФИ
            ws_child['D'+str(Find_id(callback.message))] = fio[0] #Ф
            ws_child['E' + str(Find_id(callback.message))] = fio[1] #И
        else: # если у нас ФИО
            ws_child['D' + str(Find_id(callback.message))] = fio[0] #Ф
            ws_child['E' + str(Find_id(callback.message))] = fio[1] #И
            ws_child['F' + str(Find_id(callback.message))] = fio[2] #О
        wb_child.save(fn_child) # СОХРАНЯЕМ
        wb_child.close() # ЗАКРЫВАЕМ
        bot.send_message(callback.message.chat.id, "Рад знакомству!!!🥳")
        Phone1(callback.message) # снова рассказываем пользователю о возможностях бота и даём кнопочки викторина/сертификат
    elif callback.data == 'нет, это не моё имя': # пиздюк запорол ввод имени
        bot.send_message(callback.message.chat.id, "Кажется, возникла какая то ошибка😅\n"
                                                   "попробуй ввести своё имя повторно\n"
                                                   "формат для ввода: Фамилия Имя Отчество\n\n"
                                                   "я постараюсь быть внимательнее")
        bot.register_next_step_handler(callback.message, SayMyName2) # даём пиздюку возможность снова ввести имя и возвращаемся в функцию проверки имени

    elif callback.data == "true":
        cat = int(ws_child["R" + str(Find_id(callback.message))].value)
        ws_child[ans[cat] + str(Find_id(callback.message))] = int(ws_child[ans[cat] + str(Find_id(callback.message))].value) + 1
        if (int(ws_child['R' + str(Find_id(callback.message))].value) * 6 + int(ws_child['S' + str(Find_id(callback.message))].value)) < len(questions):
            opros(callback.message)
        else:
            #bot.send_message(callback.message.chat.id, "Умничка")
            Rezultat(callback.message)
    elif callback.data == "false":
        if (int(ws_child['R' + str(Find_id(callback.message))].value) * 6 + int(ws_child['S' + str(Find_id(callback.message))].value)) < len(questions):
            opros(callback.message)
        else:
            #bot.send_message(callback.message.chat.id, "Умничка")
            Rezultat(callback.message)

@bot.message_handler(content_types=['text'])
def check_login(message):
    if ws_admin['D2'].value == message.text:
      #bot.send_message(message.chat.id, "Введите пароль")
      bot.register_next_step_handler(message, check_password)


    elif ws_admin['A2'].value == message.chat.id:
        if message.text == "Посмотреть количество участницков":
            rows = int(ws_child.max_row) - 1
            bot.send_message(message.chat.id, f"Всего записано детей: {rows}")
        if message.text == "Количество доступных мест":
            rows = int(ws_child.max_row) - 1
            number = ws_admin['G2'].value
            bot.send_message(message.chat.id, f"Всего доступно мест: {number-rows}")
        elif message.text == "Добавить места":
            rows = int(ws_child.max_row) - 1
            number = ws_admin['G2'].value
            bot.send_message(message.chat.id, f"Всего детей записано: {rows}\n"
                            f"Количество доступных мест {number-rows}\n"
                            f"Введите необходимый размер комнаты")
            bot.register_next_step_handler(message, add_specify_children)
        elif message.text == "Уменьшить места":
            rows = int(ws_child.max_row) - 1
            number = ws_admin['G2'].value
            bot.send_message(message.chat.id, f"Всего детей записано: {rows}\n"
                            f"Количество доступных мест: {number-rows}\n"
                            f"Введите сколько мест нужно убрать из комнаты")
            bot.register_next_step_handler(message, del_specify_children)
        elif message.text == "Название мастер класса":
            name = ws_admin['F2'].value
            bot.send_message(message.chat.id, f"Название мастер класса: {name}")
        elif message.text == "Указать название мастеркласса":
            bot.send_message(message.chat.id, f"Введите название мастер класса")
            bot.register_next_step_handler(message, rename_mk)
        elif message.text == "Введенный возраст детей":
            name = ws_admin['H2'].value
            bot.send_message(message.chat.id, f"Сейчас возраст детей записан как: {name}")
        elif message.text == "Указать возраст детей":
            bot.send_message(message.chat.id, f"Введите возраст")
            bot.register_next_step_handler(message, age_child)
        elif message.text == "Введенная школа":
            name = ws_admin['I2'].value
            bot.send_message(message.chat.id, f"Сейчас школа записана как: {name}")
        elif message.text == "Указать школу":
            bot.send_message(message.chat.id, f"Укажите школу")
            bot.register_next_step_handler(message, name_school)
        elif message.text == "Указать мероприятие":
            bot.send_message(message.chat.id, f"Укажите мероприятие (mk/ank)\n"
                                            f"mk - Дети будут проходить мастер-класс"
                                            f"ank - Дети будут проходить анкетирование")
            bot.register_next_step_handler(message, meropriyatie)
    else:
        pass #bot.send_message(message.chat.id, "неверно")


bot.infinity_polling()
